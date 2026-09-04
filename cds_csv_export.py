"""
CDS Customs Entry Worksheet Excel export.

Generates an .xlsx that matches the CDS upload template used by customs software,
with 4 fixed header rows followed by consolidated item data rows.
Only populates the fields that are actually completed in normal workflow.
"""
import io
import os
import warnings
import unicodedata
from typing import Dict, List, Optional
from openpyxl import Workbook, load_workbook
from consolidation import group_by_commodity_code, consolidate_items
from countries import normalize_country_iso
from hmrc_api import format_hs_for_sheet, resolve_hmrc_data

__all__ = [
    "CDS_MAX_ITEMS_PER_ENTRY",
    "create_cds_excel",
    "create_cds_excel_parts",
    "cds_export_download_meta",
    "chunk_cds_rows",
]

# Path to the original CDS template bundled with this tool.
# Using the template preserves all merged cells, column widths, and borders
# required by the CDS upload system.
TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              'templates', 'CDS-Customs-Entry-Worksheet-template.xlsx')
FIRST_DATA_ROW = 5  # Rows 1-4 are header rows in the CDS template


# ---------------------------------------------------------------------------
# Fixed header rows (rows 1-4) — must match CDS upload template exactly
# ---------------------------------------------------------------------------
HEADER_ROW_1 = [
    'Update Type', 'Entry Type', 'Tmode', 'Cons Ref', '', '', '', '',
    'Type', 'Commodity', 'Entry Version', 'Amalgamate Y/N', '', '',
    'SHCPDC', 'SADH_ALL_PREVIOUS_DOCUMENTS', '',
    'SHCAIT', 'SADH_ALL_ADD_INFO_TEXT', '', '', '', '',
    'SHCDOC', 'SADH_ALL_DOCUMENTS', '', '', '', '', '', '', '',
    'SHCTAX', 'SADH_ALL_TAX', '', '', '', '', '',
    'SHCIPA', 'SADH_ALL_PROCEDURE_ADDITIONAL', '',
    'SHCSUP', 'SADH_ALL_SUPPLEMENTARY_CODES', '',
]

HEADER_ROW_2 = [
    '9', 'H1', '1', '', '', '', '', '',
    'I', 'CDI10', '1', 'N', '', '',
    'CNPDCCLASS$$', 'CNPDCTYPE$$', 'PREV_DOC_REF',
    'CNAISTMT$$', 'AI_STMT_TXT', '', '', '', '',
    'CNDOCCODE$$', 'CNDOCSTAT$$', 'DOC_REF', '', '', '', '', '', '',
    'TAXTYPE$$', 'BASE_AMT_DC', 'SHQTYCODE$$', 'MOP_CODE', '', '', '',
    'SHCPCADD$$', '', '',
    'COMMODITY_S2$$', '', '',
]


def _make_header_row_2(is_export: bool) -> list:
    """Return HEADER_ROW_2 with direction-specific values.

    Export: Entry Type=B1, Tmode=2, Type=E, Commodity=CDE08
    Import: Entry Type=H1, Tmode=2, Type=I, Commodity=CDI10
    """
    row = list(HEADER_ROW_2)          # copy the template
    if is_export:
        row[1] = 'B1'                 # Entry Type
        row[8] = 'E'                  # Type
        row[9] = 'CDE08'              # Commodity
    else:
        row[1] = 'H1'
        row[8] = 'I'
        row[9] = 'CDI10'
    row[2] = '2'                      # Tmode — always Air
    return row

HEADER_ROW_3 = [
    '', '', '', '', '', '', '', '', '', '', '', '', '', '',
    'Previous Document', '', '',
    'AI Statement 1', '', 'AI Statement 2', '', 'AI Statement 3', '',
    'Document Code 1', '', '', 'Document Code 2', '', '', 'Document Code 3', '', '',
    'Tax Line 1', '', '', '', 'Tax Line 2', '', '', '',
    'Procedure Additionals', '', '',
    'Supplementary Commodity Codes', '', '',
]

HEADER_ROW_4 = [
    'Commodity', 'Supp.Units', 'Stats Value (Exports)', 'Item Value (Imports)',
    'Goods Description', 'Nett Weight', 'CPC', 'Gross Weight', 'Quota',
    'Package', 'Package Count', 'Country Orig', 'Prefseg1', 'Prefseg2',
    'Prv Doc Class', 'Prv Doc Type', 'Prv Doc Reference',
    'AI Code', 'AI Text', 'AI Code', 'AI Text', 'AI Code', 'AI Text',
    'Doc Code', 'Doc Status', 'Doc Reference',
    'Doc Code', 'Doc Status', 'Doc Reference',
    'Doc Code', 'Doc Status', 'Doc Reference',
    'Tax Type', 'Tax Rate Amount', 'Tax Rate Qty', 'MOP',
    'Tax Type', 'Tax Rate Amount', 'Tax Rate Qty', 'MOP',
    'Proc Add 1', 'Proc Add2', 'Proc Add3',
    'Supp 1', 'Supp2', 'Supp3',
]

NUM_COLUMNS = len(HEADER_ROW_4)          # 46
TOTAL_DATA_ROWS = 200                    # Pad to this many data rows (fallback only)
# CDS declaration item limit (DE 1/9 etc.) — split worksheets beyond this.
CDS_MAX_ITEMS_PER_ENTRY = 99


def _strip_non_ascii(text: str) -> str:
    """Replace or remove characters above ASCII 127 from a string.

    CDS upload rejects files containing high-ASCII / Unicode characters in
    goods descriptions.  Common replacements are applied first (en/em dashes,
    curly quotes, accented letters, etc.) and any remaining non-ASCII chars are
    dropped.
    """
    if not text:
        return text
    # Manual replacements for the most common problem characters
    _REPLACEMENTS = {
        '\u2013': '-',   # en dash
        '\u2014': '-',   # em dash
        '\u2018': "'",   # left single quote
        '\u2019': "'",   # right single quote
        '\u201c': '"',   # left double quote
        '\u201d': '"',   # right double quote
        '\u2026': '...',  # ellipsis
        '\u00a0': ' ',   # non-breaking space
        '\u00b0': 'deg', # degree sign
        '\u00d7': 'x',   # multiplication sign
        '\u00b1': '+/-', # plus-minus
        '\u00bd': '1/2', # half
        '\u00bc': '1/4', # quarter
        '\u00be': '3/4', # three-quarters
        '\u2122': '',    # trademark
        '\u00ae': '',    # registered trademark
        '\u00a9': '',    # copyright
        '\u00df': 'ss',  # sharp s (ß)
    }
    for ch, replacement in _REPLACEMENTS.items():
        text = text.replace(ch, replacement)
    # Decompose accented chars (é → e + combining accent) then strip non-ASCII
    text = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in text if ord(c) < 128)


def _safe_float(value, default=0.0):
    """Safely convert a value to float."""
    if value in ('', None):
        return default
    try:
        return float(str(value).replace(',', '').replace(' ', ''))
    except (ValueError, TypeError):
        return default


def _build_cds_data_rows(
    items: List[Dict],
    direction: str = 'export',
    hmrc_data: Optional[Dict[str, Dict]] = None,
    metadata: Optional[Dict] = None,
    consolidate: bool = True,
) -> List[List]:
    """Build CDS worksheet data rows (one list per declaration line)."""
    hmrc_data = hmrc_data or {}
    metadata = metadata or {}
    is_export = direction.lower() == 'export'

    package_type = metadata.get('package_type', 'PK')
    package_count = metadata.get('number_of_packages', '')
    invoice_ref = (
        metadata.get('invoice_number')
        or metadata.get('document_ref')
        or metadata.get('reference', '')
    )

    if consolidate:
        grouped = group_by_commodity_code(items)
    else:
        grouped = {f'item_{idx}': [item] for idx, item in enumerate(items)}

    data_rows: List[List] = []
    for code_key, group_items in grouped.items():
        consolidated = consolidate_items(group_items)

        if consolidate:
            commodity_code = code_key
            if commodity_code.startswith('__BLANK_'):
                commodity_code = ''
        else:
            commodity_code = group_items[0].get('commodity_code', '')

        descriptions: List[str] = []
        seen_descs: set = set()
        for it in group_items:
            d = it.get('description', '').strip()
            if d and d not in seen_descs:
                descriptions.append(d)
                seen_descs.add(d)
        description = _strip_non_ascii('; '.join(descriptions))

        total_value = (
            round(consolidated['total_value'], 2) if consolidated['total_value'] else ''
        )
        net_weight = (
            round(consolidated['total_net_weight'], 3)
            if consolidated['total_net_weight'] is not None
            and consolidated['total_net_weight'] != 0
            else ''
        )
        gross_weight = (
            round(consolidated['total_net_weight'] * 1.10, 3)
            if consolidated['total_net_weight'] is not None
            and consolidated['total_net_weight'] != 0
            else ''
        )

        countries = consolidated.get('countries_of_origin', [])
        country_orig = normalize_country_iso(countries[0] if countries else '')

        # Sheet shows CN8 for export; HMRC cache may be keyed by 8 or 10 digits.
        hmrc_info = resolve_hmrc_data(hmrc_data, commodity_code)
        sheet_code = format_hs_for_sheet(commodity_code, direction)

        supp_units = ''
        supp_raw = hmrc_info.get('supplementary_units', '')
        if supp_raw and supp_raw != 'Not required':
            supp_units = (
                int(consolidated['total_quantity']) if consolidated['total_quantity'] else ''
            )

        doc_codes_dict = (
            hmrc_info.get('selected_document_codes') or hmrc_info.get('document_codes', {})
        )
        doc_list = list(doc_codes_dict.keys())[:3]

        row = [''] * NUM_COLUMNS
        row[0] = sheet_code
        row[1] = supp_units
        if is_export:
            row[2] = total_value
        else:
            row[3] = total_value
        row[4] = description
        row[5] = net_weight
        row[7] = gross_weight
        row[9] = package_type if package_type else 'PK'
        row[10] = package_count
        row[11] = country_orig
        row[14] = 'Z'
        row[15] = '380'
        if invoice_ref:
            row[16] = invoice_ref
        if len(doc_list) >= 1:
            row[23] = doc_list[0]
        if len(doc_list) >= 2:
            row[26] = doc_list[1]
        if len(doc_list) >= 3:
            row[29] = doc_list[2]
        data_rows.append(row)

    return data_rows


def _write_cds_workbook(data_rows: List[List], is_export: bool) -> io.BytesIO:
    """Write one CDS template workbook for the given data rows."""
    try:
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        ws['B2'] = 'B1' if is_export else 'H1'
        ws['C2'] = 2
        ws['I2'] = 'E' if is_export else 'I'
        ws['J2'] = 'CDE08' if is_export else 'CDI10'

        max_row = max(ws.max_row, FIRST_DATA_ROW + len(data_rows))
        for r in range(FIRST_DATA_ROW, max_row + 1):
            for c in range(1, ws.max_column + 1):
                try:
                    ws.cell(row=r, column=c).value = None
                except AttributeError:
                    pass

    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        for header_row in [HEADER_ROW_1, _make_header_row_2(is_export), HEADER_ROW_3, HEADER_ROW_4]:
            ws.append(header_row)

    for row_idx, row_data in enumerate(data_rows, start=FIRST_DATA_ROW):
        for col_idx, val in enumerate(row_data, start=1):
            if val not in ('', None):
                try:
                    ws.cell(row=row_idx, column=col_idx).value = val
                except AttributeError:
                    pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def chunk_cds_rows(
    data_rows: List[List],
    max_items: int = CDS_MAX_ITEMS_PER_ENTRY,
) -> List[List[List]]:
    """Split declaration lines into CDS-sized chunks (default 99)."""
    if max_items <= 0:
        return [data_rows]
    if not data_rows:
        return [[]]
    return [
        data_rows[i:i + max_items]
        for i in range(0, len(data_rows), max_items)
    ]


def create_cds_excel_parts(
    items: List[Dict],
    direction: str = 'export',
    hmrc_data: Optional[Dict[str, Dict]] = None,
    metadata: Optional[Dict] = None,
    consolidate: bool = True,
    max_items_per_file: int = CDS_MAX_ITEMS_PER_ENTRY,
) -> List[Dict]:
    """
    Create one or more CDS worksheets, splitting at the CDS item limit.

    Returns a list of dicts:
      {part, parts, item_count, start_item, end_item, buffer}
    """
    is_export = direction.lower() == 'export'
    data_rows = _build_cds_data_rows(
        items, direction=direction, hmrc_data=hmrc_data,
        metadata=metadata, consolidate=consolidate,
    )
    chunks = chunk_cds_rows(data_rows, max_items=max_items_per_file)
    parts: List[Dict] = []
    offset = 0
    for idx, chunk in enumerate(chunks, start=1):
        parts.append({
            'part': idx,
            'parts': len(chunks),
            'item_count': len(chunk),
            'start_item': offset + 1 if chunk else 0,
            'end_item': offset + len(chunk),
            'buffer': _write_cds_workbook(chunk, is_export),
        })
        offset += len(chunk)
    return parts


def create_cds_excel(
    items: List[Dict],
    direction: str = 'export',
    hmrc_data: Optional[Dict[str, Dict]] = None,
    metadata: Optional[Dict] = None,
    consolidate: bool = True,
    max_items_per_file: int = CDS_MAX_ITEMS_PER_ENTRY,
) -> io.BytesIO:
    """
    Create a CDS Customs Entry Worksheet as an Excel (.xlsx) file.

    If consolidated declaration lines exceed ``max_items_per_file`` (CDS limit
    99), returns a ZIP containing part files instead of a single workbook.

    Loads the original CDS template (preserving merged cells, column widths and
    borders) then writes direction-specific header values and item data rows.

    Args:
        items:     Parsed line items. If consolidate=True (default), will be
                   consolidated by HS code. If consolidate=False, items are
                   assumed to already be consolidated (or no consolidation needed).
        direction: 'export' or 'import'.
        hmrc_data: HMRC tariff lookup results keyed by commodity code.
        metadata:  Invoice metadata (number_of_packages, package_type, cpc_code...).
        consolidate: If True, consolidate raw items by commodity code.
                     If False, items are already consolidated/finalized.
        max_items_per_file: CDS max declaration lines per worksheet (default 99).

    Returns:
        BytesIO containing .xlsx (one part) or .zip (multiple parts).
    """
    parts = create_cds_excel_parts(
        items=items,
        direction=direction,
        hmrc_data=hmrc_data,
        metadata=metadata,
        consolidate=consolidate,
        max_items_per_file=max_items_per_file,
    )
    if len(parts) == 1:
        return parts[0]['buffer']

    import zipfile
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for p in parts:
            name = (
                f"CDS-Customs-Entry-Worksheet-FCL-"
                f"part{p['part']}of{p['parts']}-"
                f"items{p['start_item']}-{p['end_item']}.xlsx"
            )
            zf.writestr(name, p['buffer'].getvalue())
    zip_buf.seek(0)
    return zip_buf


def cds_export_download_meta(parts: List[Dict], base_name: str) -> Dict:
    """Filename / mime / caption for a create_cds_excel_parts result."""
    if len(parts) <= 1:
        return {
            'base_name': base_name,
            'file_name': f"{base_name}.xlsx",
            'mime': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'label': '📋 Download FCL Excel Sheet',
            'caption': None,
            'data': parts[0]['buffer'].getvalue() if parts else b'',
            'parts': parts,
        }
    import zipfile
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for p in parts:
            name = (
                f"{base_name}-part{p['part']}of{p['parts']}-"
                f"items{p['start_item']}-{p['end_item']}.xlsx"
            )
            zf.writestr(name, p['buffer'].getvalue())
    zip_buf.seek(0)
    total = sum(p['item_count'] for p in parts)
    return {
        'base_name': base_name,
        'file_name': f"{base_name}-split-{len(parts)}parts.zip",
        'mime': 'application/zip',
        'label': f'📋 Download FCL Excel ZIP ({len(parts)} files)',
        'caption': (
            f"CDS allows max {CDS_MAX_ITEMS_PER_ENTRY} items per entry — "
            f"split {total} lines into {len(parts)} worksheets of up to "
            f"{CDS_MAX_ITEMS_PER_ENTRY}."
        ),
        'data': zip_buf.getvalue(),
        'parts': parts,
    }
