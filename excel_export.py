"""
Enhanced Excel export with HMRC data integration
Creates comprehensive exports with document codes in separate columns
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from typing import Dict, List, Optional
import io


def _safe_float(value, default=0.0):
    """Safely convert a value to float, handling empty strings and None"""
    if value in ['', None]:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _calculate_gross_weight(net_weight: float, items: List[Dict]) -> float:
    """
    Calculate gross weight from net weight using intelligent estimation.
    
    Args:
        net_weight: Total net weight
        items: List of items (may contain gross_weight field)
        
    Returns:
        Estimated or actual gross weight
    """
    # First, check if any items have actual gross weight
    actual_gross = []
    for item in items:
        gw = item.get('gross_weight')
        if gw:
            actual_gross.append(_safe_float(gw))
    
    if actual_gross:
        # If we have real gross weights, sum them
        return sum(actual_gross)
    
    # Otherwise, estimate from net weight
    if net_weight and net_weight > 0:
        # Standard packaging allowance: 10% for general goods
        # This can be adjusted based on commodity type if needed
        return round(net_weight * 1.10, 3)
    
    return 0.0


def _summarize_requirement(requirement: str) -> str:
    """Intelligently shorten requirement text while keeping essential info"""
    if not requirement or len(requirement) <= 50:
        return requirement
    
    import re
    
    # Remove all parenthetical content containing "Regulation" 
    # Strategy: find and remove the outermost parentheses that contain "regulation"
    shortened = requirement
    max_iterations = 10  # Prevent infinite loops
    iteration = 0
    
    while '(' in shortened and iteration < max_iterations:
        iteration += 1
        # Find opening parenthesis
        start = shortened.find('(')
        if start == -1:
            break
        
        # Find matching closing parenthesis
        depth = 0
        end = -1
        for i in range(start, len(shortened)):
            if shortened[i] == '(':
                depth += 1
            elif shortened[i] == ')':
                depth -= 1
                if depth == 0:
                    end = i
                    break
        
        if end == -1:
            break
        
        # Check if this section contains "regulation"
        section = shortened[start:end+1]
        if 'regulation' in section.lower():
            # Remove this parenthetical section
            shortened = shortened[:start] + shortened[end+1:]
        else:
            # No regulation in this section, no more to process
            break
    
    # Common phrase replacements
    replacements = {
        'Export licence: Dual use export authorisation': 'Dual Use Export Licence',
        'Export licence: ': '',
        'Import licence: ': '',
        'Certificate of origin': 'Origin Certificate',
        'Phytosanitary certificate': 'Phytosanitary Cert',
        'Veterinary certificate': 'Veterinary Cert',
        'Health certificate': 'Health Cert',
        'Conformity certificate': 'Conformity Cert',
        'Commercial invoice': 'Comm. Invoice',
        'Packing list': 'Packing List',
        ' required': '',
        ' is required': '',
    }
    
    for old, new in replacements.items():
        shortened = shortened.replace(old, new)
    
    # Trim whitespace and limit length
    shortened = shortened.strip()
    if len(shortened) > 60:
        shortened = shortened[:57] + '...'
    
    return shortened


def create_comprehensive_export(
    items: List[Dict],
    hmrc_data: Optional[Dict[str, Dict]] = None,
    direction: str = 'export',
    country: str = '',
    consolidate: bool = False,
    metadata: Optional[Dict] = None
) -> io.BytesIO:
    """
    Create comprehensive Excel export with all invoice and HMRC data.
    
    Args:
        items: List of parsed invoice items
        hmrc_data: HMRC tariff data keyed by commodity code
        direction: 'import' or 'export'
        country: Destination/origin country
        consolidate: Whether to consolidate items by commodity code
        metadata: Invoice metadata (incoterms, CPC, etc.)
        
    Returns:
        BytesIO object containing the Excel file
    """
    metadata = metadata or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Customs Declaration"
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Build column headers based on direction
    if direction == 'export':
        base_headers = [
            'Commodity Code',
            'Description',
            'Quantity',
            'UOM',
            'Value (£)',
            'Currency',
            'Net Weight (kg)',
            'Gross Weight (kg)',
            'Country of Origin'
        ]
        
        # Add HMRC columns if data available
        if hmrc_data:
            base_headers.extend([
                'Supplementary Units',
                'Doc Code 1',
                'Doc Code 1 - Requirement',
                'Doc Code 2',
                'Doc Code 2 - Requirement',
                'Doc Code 3',
                'Doc Code 3 - Requirement',
                'Doc Code 4',
                'Doc Code 4 - Requirement',
                'Additional Notes'
            ])
    else:  # import
        base_headers = [
            'Commodity Code',
            'Description',
            'Quantity',
            'UOM',
            'Value (£)',
            'Currency',
            'Net Weight (kg)',
            'Gross Weight (kg)',
            'Country of Origin'
        ]
        
        if hmrc_data:
            base_headers.extend([
                'CPC Code',
                'Valuation Method',
                'VAT Rate',
                'Third Country Duty',
                'Supplementary Units',
                'Doc Code 1',
                'Doc Code 1 - Requirement',
                'Doc Code 2',
                'Doc Code 2 - Requirement',
                'Doc Code 3',
                'Doc Code 3 - Requirement',
                'Additional Notes'
            ])
    
    # Write headers
    for col_num, header in enumerate(base_headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Process items
    if consolidate and hmrc_data:
        # Group by commodity code
        grouped = {}
        for item in items:
            code = item.get('commodity_code', 'Unknown')
            if code not in grouped:
                grouped[code] = []
            grouped[code].append(item)
        
        # Write consolidated rows
        row_num = 2
        for commodity_code, group_items in grouped.items():
            # Consolidate quantities and values - handle empty strings and None
            total_qty = sum(_safe_float(item.get('quantity')) for item in group_items)
            total_value = sum(_safe_float(item.get('total_value') or item.get('value')) for item in group_items)
            total_weight = sum(_safe_float(item.get('net_weight')) for item in group_items)
            
            # Get first item for reference data
            first_item = group_items[0]
            
            # Get all unique countries of origin
            countries = set()
            for item in group_items:
                coo = item.get('country_of_origin', '')
                if coo:
                    countries.add(coo)
            
            # Determine description: prefer invoice item descriptions for customs declarations
            # HMRC tariff descriptions can be too generic (e.g., "other")
            description = ''
            
            # Collect unique descriptions from all items in this group
            unique_descriptions = []
            for item in group_items:
                desc = item.get('description', '').strip()
                if desc and desc not in unique_descriptions:
                    unique_descriptions.append(desc)
            
            if len(unique_descriptions) == 1:
                # Single description - use it
                description = unique_descriptions[0]
            elif len(unique_descriptions) > 1:
                # Multiple descriptions - combine them, truncate if too long
                combined = '; '.join(unique_descriptions)
                description = combined[:100] + '...' if len(combined) > 100 else combined
            
            # Only fall back to HMRC description if no invoice descriptions exist
            if not description:
                if hmrc_data and commodity_code in hmrc_data:
                    hmrc_info = hmrc_data[commodity_code]
                    if 'error' not in hmrc_info:
                        description = hmrc_info.get('description', '')
            
            # Final fallback
            if not description:
                description = first_item.get('description', '')
            
            # Collect review notes from any item in the group
            review_notes = []
            for gi in group_items:
                rn = gi.get('review_notes', '')
                if rn and rn not in review_notes:
                    review_notes.append(rn)

            row_data = {
                'Description': description,
                'Commodity Code': commodity_code,
                'Quantity': total_qty,
                'UOM': first_item.get('uom', 'ea'),
                'Value': total_value,
                'Currency': first_item.get('currency', 'GBP'),
                'Weight': total_weight,
                'Gross Weight': _calculate_gross_weight(total_weight, group_items),
                'Country': ', '.join(countries) if countries else first_item.get('country_of_origin', ''),
                '_review_notes': '; '.join(review_notes) if review_notes else ''
            }
            
            _write_row(ws, row_num, row_data, hmrc_data.get(commodity_code) if hmrc_data else None, 
                      direction, thin_border, metadata)
            row_num += 1
    else:
        # Write individual items
        row_num = 2
        for item in items:
            commodity_code = item.get('commodity_code', '')
            
            # Calculate gross weight if missing
            net_wt = _safe_float(item.get('net_weight'))
            gross_wt = item.get('gross_weight', '')
            if not gross_wt and net_wt > 0:
                # Estimate gross as net + 10% packaging
                gross_wt = round(net_wt * 1.10, 3)
            
            row_data = {
                'Description': item.get('description', ''),
                'Commodity Code': commodity_code,
                'Quantity': item.get('quantity', ''),
                'UOM': item.get('uom', 'ea'),
                'Value': item.get('total_value') or item.get('value', ''),
                'Currency': item.get('currency', 'GBP'),
                'Weight': net_wt,
                'Gross Weight': gross_wt,
                'Country': item.get('country_of_origin', ''),
                '_review_notes': item.get('review_notes', '')
            }
            
            _write_row(ws, row_num, row_data, hmrc_data.get(commodity_code) if hmrc_data else None,
                      direction, thin_border, metadata)
            row_num += 1
    
    # Adjust column widths
    column_widths = {
        'A': 18,  # Commodity Code
        'B': 40,  # Description
        'C': 12,  # Quantity
        'D': 8,   # UOM
        'E': 12,  # Value
        'F': 10,  # Currency
        'G': 15,  # Net Weight
        'H': 15,  # Gross Weight
        'I': 20,  # Country
    }
    
    # HMRC columns (start at J for both directions)
    if hmrc_data:
        if direction == 'export':
            column_widths.update({
                'J': 20,   # Supp Units
                'K': 12,   # Doc Code 1
                'L': 50,   # Doc Code 1 Req
                'M': 12,   # Doc Code 2
                'N': 50,   # Doc Code 2 Req
                'O': 12,   # Doc Code 3
                'P': 50,   # Doc Code 3 Req
                'Q': 12,   # Doc Code 4
                'R': 50,   # Doc Code 4 Req
                'S': 40,   # Notes
            })
        else:
            column_widths.update({
                'J': 12,   # CPC Code
                'K': 15,   # Valuation Method
                'L': 15,   # VAT
                'M': 20,   # Duty
                'N': 20,   # Supp Units
                'O': 12,   # Doc Code 1
                'P': 50,   # Doc Code 1 Req
                'Q': 12,   # Doc Code 2
                'R': 50,   # Doc Code 2 Req
                'S': 12,   # Doc Code 3
                'T': 50,   # Doc Code 3 Req
                'U': 40,   # Notes
            })
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    return excel_bytes


def _write_row(ws, row_num: int, row_data: Dict, hmrc_info: Optional[Dict], 
               direction: str, border,metadata: Optional[Dict] = None):
    """Write a single row with item and HMRC data"""
    
    metadata = metadata or {}
    
    # Write base columns (Commodity Code first, then Description)
    cc_cell = ws.cell(row=row_num, column=1)
    cc_cell.value = row_data.get('Commodity Code', '')

    # Highlight and add comment if review notes exist
    review_notes = row_data.get('_review_notes', '')
    if review_notes:
        cc_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        cc_cell.comment = Comment(review_notes, 'Auto-Review')

    ws.cell(row=row_num, column=2).value = row_data.get('Description', '')
    ws.cell(row=row_num, column=3).value = row_data.get('Quantity', '')
    ws.cell(row=row_num, column=4).value = row_data.get('UOM', '')
    ws.cell(row=row_num, column=5).value = row_data.get('Value', '')
    ws.cell(row=row_num, column=6).value = row_data.get('Currency', '')
    ws.cell(row=row_num, column=7).value = row_data.get('Weight', '')
    
    col_num = 8
    
    # Gross weight column (both import and export)
    ws.cell(row=row_num, column=col_num).value = row_data.get('Gross Weight', '')
    col_num += 1
    
    ws.cell(row=row_num, column=col_num).value = row_data.get('Country', '')
    col_num += 1
    
    # Start of HMRC columns
    
    # Add HMRC data if available
    if hmrc_info and 'error' not in hmrc_info:
        if direction == 'export':
            # Supplementary units
            supp_units = hmrc_info.get('supplementary_units') or 'N/A'
            ws.cell(row=row_num, column=col_num).value = supp_units
            col_num += 1
            
            # Document codes (up to 4)
            doc_codes = hmrc_info.get('document_codes', {})
            doc_items = list(doc_codes.items())
            
            # If no document codes at all, write N/A in first code column
            if not doc_codes:
                ws.cell(row=row_num, column=col_num).value = 'N/A'
                ws.cell(row=row_num, column=col_num + 1).value = 'No document codes required'
                col_num += 2
                # Leave remaining columns empty
                for i in range(3):
                    ws.cell(row=row_num, column=col_num).value = ''
                    ws.cell(row=row_num, column=col_num + 1).value = ''
                    col_num += 2
            else:
                # Write document codes
                for i in range(4):
                    if i < len(doc_items):
                        code, requirement = doc_items[i]
                        ws.cell(row=row_num, column=col_num).value = code
                        ws.cell(row=row_num, column=col_num + 1).value = _summarize_requirement(requirement)
                    else:
                        ws.cell(row=row_num, column=col_num).value = ''
                        ws.cell(row=row_num, column=col_num + 1).value = ''
                    col_num += 2
            
            # Additional notes
            notes = []
            if len(doc_codes) > 4:
                notes.append(f"Additional codes: {', '.join(list(doc_codes.keys())[4:])}")
            
            add_codes = hmrc_info.get('additional_codes', [])
            if add_codes:
                codes_str = ', '.join([ac.get('code', '') for ac in add_codes[:3]])
                notes.append(f"Additional codes: {codes_str}")
            
            ws.cell(row=row_num, column=col_num).value = '; '.join(notes) if notes else ''
            
        else:  # import
            # CPC Code (Customs Procedure Code)
            cpc_code = metadata.get('cpc_code', '4000')
            ws.cell(row=row_num, column=col_num).value = cpc_code
            col_num += 1
            
            # Valuation Method
            val_method = metadata.get('valuation_method', '1')
            ws.cell(row=row_num, column=col_num).value = val_method
            col_num += 1
            
            # VAT Rate
            ws.cell(row=row_num, column=col_num).value = hmrc_info.get('vat_rate', 'N/A')
            col_num += 1
            
            # Third Country Duty
            ws.cell(row=row_num, column=col_num).value = hmrc_info.get('third_country_duty', 'N/A')
            col_num += 1
            
            # Supplementary units
            supp_units = hmrc_info.get('supplementary_units') or 'N/A'
            ws.cell(row=row_num, column=col_num).value = supp_units
            col_num += 1
            
            # Document codes (up to 3 for import)
            doc_codes = hmrc_info.get('document_codes', {})
            doc_items = list(doc_codes.items())
            
            # If no document codes at all, write N/A in first code column
            if not doc_codes:
                ws.cell(row=row_num, column=col_num).value = 'N/A'
                ws.cell(row=row_num, column=col_num + 1).value = 'No document codes required'
                col_num += 2
                # Leave remaining columns empty
                for i in range(2):
                    ws.cell(row=row_num, column=col_num).value = ''
                    ws.cell(row=row_num, column=col_num + 1).value = ''
                    col_num += 2
            else:
                # Write document codes
                for i in range(3):
                    if i < len(doc_items):
                        code, requirement = doc_items[i]
                        ws.cell(row=row_num, column=col_num).value = code
                        ws.cell(row=row_num, column=col_num + 1).value = _summarize_requirement(requirement)
                    else:
                        ws.cell(row=row_num, column=col_num).value = ''
                        ws.cell(row=row_num, column=col_num + 1).value = ''
                    col_num += 2
            
            # Additional notes
            notes = []
            if len(doc_codes) > 3:
                notes.append(f"More doc codes: {', '.join(list(doc_codes.keys())[3:])}")
            
            ws.cell(row=row_num, column=col_num).value = '; '.join(notes) if notes else ''
    
    # Apply borders to all cells in row
    for col in range(1, col_num + 1):
        ws.cell(row=row_num, column=col).border = border
